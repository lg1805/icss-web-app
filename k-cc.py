from flask import Flask, request, render_template, send_file
import pandas as pd
import os
import joblib
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads/processed/'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Paths for model
MODEL_PATH = r"D:\Lakshya\Project\ICSS_VSCode\PROJECT\ICSS Web App\random_forest_model.pkl"

# Load trained model
def load_model():
    try:
        model_data = joblib.load(MODEL_PATH)
        if isinstance(model_data, tuple):
            model, vectorizer = model_data
        else:
            model = model_data
            vectorizer = None
        
        if not hasattr(model, 'predict'):
            raise ValueError("Loaded model does not have a 'predict' method")
        print("Model loaded successfully")
        return model, vectorizer
    except Exception as e:
        print(f"Error loading model: {e}")
        return None, None

model, vectorizer = load_model()

@app.route('/')
def index():
    return render_template('front1.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'complaint_file' not in request.files:
        return "No complaint_file part", 400
    
    file = request.files['complaint_file']
    if file.filename == '':
        return "No selected file", 400
    
    if file:
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        
        try:
            df = pd.read_excel(filepath)
        except Exception as e:
            return f"Error reading file: {e}", 400
        
        if 'Observation' not in df.columns or 'Incident Status' not in df.columns:
            return "Error: Required columns are missing in the uploaded file.", 400
        
        # Segregate SPN and Non-SPN complaints
        spn_df = df[df['Observation'].str.contains('SPN', case=False, na=False)]
        non_spn_df = df[~df['Observation'].str.contains('SPN', case=False, na=False)].copy()
        
        # Assign priority using Random Forest Model
        def predict_priority(observation):
            if model and vectorizer:
                transformed_text = vectorizer.transform([str(observation)])
                return model.predict(transformed_text)[0]
            return 'Low'
        
        if not non_spn_df.empty:
            try:
                non_spn_df['Priority'] = non_spn_df['Observation'].apply(predict_priority)
            except Exception as e:
                return f"Error during priority assignment: {e}", 400
        
        # Save processed file
        processed_filepath = os.path.join(UPLOAD_FOLDER, 'processed_' + file.filename)
        with pd.ExcelWriter(processed_filepath, engine='openpyxl') as writer:
            spn_df.to_excel(writer, sheet_name='SPN_Complaints', index=False)
            non_spn_df.to_excel(writer, sheet_name='Non_SPN_Complaints', index=False)
        
        # Load the saved Excel file for formatting
        wb = load_workbook(processed_filepath)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            
            # Find the column index for 'Incident Status'
            incident_status_col = None
            for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
                if col[0].value == 'Incident Status':
                    incident_status_col = col_idx
                    break
            
            if incident_status_col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=incident_status_col)
                    if isinstance(cell.value, str) and cell.value.lower() in ["closed", "completed"]:
                        cell.fill = green_fill
        
        wb.save(processed_filepath)
        return send_file(processed_filepath, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
