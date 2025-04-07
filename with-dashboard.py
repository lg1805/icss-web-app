from flask import Flask, request, render_template, send_file, jsonify
import pandas as pd
import os
import joblib
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import io
import base64

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads/processed/'
DATA_FOLDER = 'uploads/data/'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DATA_FOLDER, exist_ok=True)

MODEL_PATH = r"D:\Lakshya\Project\ICSS_VSCode\PROJECT\ICSS Web App\random_forest_model.pkl"

def load_model():
    try:
        model_data = joblib.load(MODEL_PATH)
        model, vectorizer = model_data if isinstance(model_data, tuple) else (model_data, None)
        if not hasattr(model, 'predict'):
            raise ValueError("Loaded model does not have a 'predict' method")
        print("Model loaded successfully")
        return model, vectorizer
    except Exception as e:
        print(f"Error loading model: {e}")
        return None, None

model, vectorizer = load_model()

@app.route('/dashboard', methods=['GET'])
def dashboard():
    past_week = datetime.now() - timedelta(days=7)
    all_files = [f for f in os.listdir(DATA_FOLDER) if f.endswith('.xlsx')]
    
    priority_data = {'High': [], 'Moderate': [], 'Low': []}
    
    for file in all_files:
        file_path = os.path.join(DATA_FOLDER, file)
        try:
            df = pd.read_excel(file_path)
            
            if 'Creation Date' in df.columns:
                df['Creation Date'] = pd.to_datetime(df['Creation Date'], errors='coerce')
                df = df[df['Creation Date'] >= past_week]
            
            if 'Observation' in df.columns and 'Priority' in df.columns:
                for _, row in df.iterrows():
                    observation, priority = row['Observation'], row['Priority']
                    if pd.notna(observation) and pd.notna(priority):
                        priority_data.setdefault(priority, []).append(observation)
        except Exception as e:
            print(f"Error processing {file}: {e}")
    
    top_complaints = {}
    for priority, observations in priority_data.items():
        top_complaints[priority] = pd.Series(observations).value_counts().nlargest(10).to_dict()
    
    # Generate a bar chart
    fig, ax = plt.subplots()
    
    categories = []
    counts = []
    
    for priority, complaints in top_complaints.items():
        for complaint, count in complaints.items():
            categories.append(f"{priority}: {complaint[:20]}...")  # Limit text length
            counts.append(count)
    
    ax.barh(categories, counts, color=['red' if 'High' in cat else 'yellow' if 'Moderate' in cat else 'green' for cat in categories])
    ax.set_xlabel("Frequency")
    ax.set_title("Top 10 Weekly Complaints")
    
    # Convert plot to image
    img = io.BytesIO()
    plt.savefig(img, format='png', bbox_inches='tight')
    img.seek(0)
    chart_url = base64.b64encode(img.getvalue()).decode()
    
    return render_template('dashboard.html', chart_url=chart_url, last_updated=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

@app.route('/highlight_status', methods=['POST'])
def highlight_status():
    file = request.files['file']
    if not file:
        return "No file uploaded", 400
    
    file_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(file_path)
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.column_letter == 'Incident Status' and isinstance(cell.value, str):
                    if cell.value.strip().lower() in ["closed", "completed"]:
                        cell.fill = green_fill
        
        wb.save(file_path)
        return send_file(file_path, as_attachment=True)
    except Exception as e:
        return str(e), 500

if __name__ == '__main__':
    app.run(debug=True)
