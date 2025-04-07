from flask import Flask, render_template, request, send_file
import pandas as pd
import os
from openpyxl import load_workbook
from flask import Flask, render_template, request, send_file
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

# Load keyword file once at startup
keyword_file = "Genset_Components_Priority_Cleaned.xlsx"
keyword_df = pd.read_excel(keyword_file, engine='openpyxl')
print("Keyword file loaded successfully.")

# Function to segregate SPN and Non-SPN complaints
def segregate_spn_nonspn(df):
    if 'Observation' not in df.columns:
        print("Error: 'Observation' column missing")
        return None, None
    
    df['Observation'] = df['Observation'].astype(str)  # Ensure it's a string
    spn_complaints = df[df['Observation'].str.contains(r'\\bspn\\b', case=False, na=False, regex=True)]
    non_spn_complaints = df[~df.index.isin(spn_complaints.index)]
    
    print(f"SPN complaints: {len(spn_complaints)}, Non-SPN complaints: {len(non_spn_complaints)}")
    return spn_complaints, non_spn_complaints

# Function to apply priority based on keyword file
def apply_priority(non_spn_df, keyword_df):
    if 'Component / System' not in keyword_df.columns or 'Priority' not in keyword_df.columns:
        print("Error: Keyword file missing required columns")
        return non_spn_df  # Return without modification if necessary columns are missing
    
    keyword_dict = dict(zip(keyword_df['Component / System'].astype(str).str.lower(), keyword_df['Priority']))
    
    def get_priority(observation):
        for key in keyword_dict:
            if key in str(observation).lower():
                return keyword_dict[key]
        return 'Low'  # Default priority if no match found
    
    non_spn_df['Priority'] = non_spn_df['Observation'].apply(get_priority)
    print("Priority applied to Non-SPN complaints.")
    return non_spn_df

@app.route('/')
def index():
    return render_template('front1.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'complaint_file' not in request.files:
        return "Missing file"
    
    complaint_file = request.files['complaint_file']
    
    if complaint_file.filename == '':
        return "No selected file"
    
    try:
        df = pd.read_excel(complaint_file, engine='openpyxl')  # Read complaint file
        print("Complaint file loaded successfully.")
        
        spn_df, non_spn_df = segregate_spn_nonspn(df)
        
        if spn_df is None and non_spn_df is None:
            return "Invalid file: Missing 'Observation' column"
        
        # Apply priority to Non-SPN complaints
        non_spn_df = apply_priority(non_spn_df, keyword_df)
        
        output_file = "Segregated_Complaints.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            spn_df.to_excel(writer, sheet_name='SPN_Complaints', index=False)
            non_spn_df.to_excel(writer, sheet_name='Non_SPN_Complaints', index=False)
        
        # Load workbook to apply formatting
        wb = load_workbook(output_file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            
            # Apply Normal Green color to 'closed' and 'completed' in 'Incident Status' column
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            
            if 'Incident Status' in df.columns:
                col_idx = df.columns.get_loc('Incident Status') + 1  # Excel column index (1-based)
                for row in range(2, ws.max_row + 1):  # Start from row 2 to skip header
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value and str(cell.value).strip().lower() in ['closed', 'completed']:
                        cell.fill = green_fill
        
        wb.save(output_file)
        print("File processing complete. Ready for download.")
        
        return send_file(output_file, as_attachment=True)
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        return f"Error processing file: {str(e)}"

if __name__ == '__main__':
    app.run(debug=True)
