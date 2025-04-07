from flask import Flask, request, render_template, send_file
import pandas as pd
import os
import joblib
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads/processed/'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Paths for model
MODEL_PATH = r"D:\Lakshya\Project\ICSS_VSCode\PROJECT\ICSS Web App\random_forest_model.pkl"

# Load trained model
def load_model():
    try:
        model_data = joblib.load(MODEL_PATH)
        if isinstance(model_data, tuple):
            model, vectorizer = model_data
        else:
            model = model_data
            vectorizer = None
        
        if not hasattr(model, 'predict'):
            raise ValueError("Loaded model does not have a 'predict' method")
        print("Model loaded successfully")
        return model, vectorizer
    except Exception as e:
        print(f"Error loading model: {e}")
        return None, None

model, vectorizer = load_model()

def standardize_date(date_str):
    possible_formats = ["%d-%m-%Y %H:%M", "%Y-%d-%m %H:%M", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M"]
    for fmt in possible_formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None

@app.route('/')
def index():
    return render_template('front1.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'complaint_file' not in request.files:
        return "No complaint_file part", 400
    
    file = request.files['complaint_file']
    if file.filename == '':
        return "No selected file", 400
    
    if file:
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        
        try:
            df = pd.read_excel(filepath)
        except Exception as e:
            return f"Error reading file: {e}", 400
        
        if 'Observation' not in df.columns or 'Incident Status' not in df.columns:
            return "Error: Required columns are missing in the uploaded file.", 400
        
        # Standardize 'Creation Date' column
        if 'Creation Date' in df.columns:
            df['Creation Date'] = df['Creation Date'].astype(str).apply(standardize_date)
            df = df.dropna(subset=['Creation Date'])
        
        # Segregate SPN and Non-SPN complaints
        spn_df = df[df['Observation'].str.contains('SPN', case=False, na=False)]
        non_spn_df = df[~df['Observation'].str.contains('SPN', case=False, na=False)].copy()
        
        # Assign priority using Random Forest Model
        def predict_priority(observation):
            if model and vectorizer:
                transformed_text = vectorizer.transform([str(observation)])
                return model.predict(transformed_text)[0]
            return 'Low'
        
        if not non_spn_df.empty:
            try:
                non_spn_df['Priority'] = non_spn_df['Observation'].apply(predict_priority)
            except Exception as e:
                return f"Error during priority assignment: {e}", 400
        
        # Ensure High > Moderate > Low priority ordering
        priority_order = {"High": 1, "Moderate": 2, "Low": 3}
        non_spn_df["Priority_Sort"] = non_spn_df["Priority"].map(priority_order)
        non_spn_df = non_spn_df.sort_values(by="Priority_Sort").drop(columns=["Priority_Sort"])
        
        # Save processed file
        processed_filepath = os.path.join(UPLOAD_FOLDER, 'processed_' + file.filename)
        with pd.ExcelWriter(processed_filepath, engine='openpyxl') as writer:
            spn_df.to_excel(writer, sheet_name='SPN_Complaints', index=False)
            non_spn_df.to_excel(writer, sheet_name='Non_SPN_Complaints', index=False)
        
        # Load the saved Excel file for formatting
        wb = load_workbook(processed_filepath)
        blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Find column indexes
            incident_status_col = None
            creation_date_col = None
            for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
                if col[0].value == 'Incident Status':
                    incident_status_col = col_idx
                if col[0].value == 'Creation Date':
                    creation_date_col = col_idx
            
            # Apply formatting
            if incident_status_col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=incident_status_col)
                    if isinstance(cell.value, str) and cell.value.lower() in ["closed", "completed"]:
                        cell.fill = green_fill
            
            if creation_date_col:
                current_date = datetime.now()
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=creation_date_col)
                    if isinstance(cell.value, datetime):
                        days_passed = (current_date - cell.value).days
                        if days_passed >= 3:
                            cell.fill = red_fill
                        elif days_passed == 2:
                            cell.fill = yellow_fill
                        elif days_passed == 1:
                            cell.fill = blue_fill
        
        wb.save(processed_filepath)
        return send_file(processed_filepath, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
