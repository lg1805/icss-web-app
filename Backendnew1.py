from flask import Flask, render_template, request, send_file
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

# Function to segregate SPN and Non-SPN complaints
def segregate_spn_nonspn(df):
    if 'Observation' not in df.columns:
        return None, None
    
    df['Observation'] = df['Observation'].astype(str)  # Ensure it's a string
    spn_complaints = df[df['Observation'].str.contains(r'\bspn\b', case=False, na=False, regex=True)]
    non_spn_complaints = df[~df.index.isin(spn_complaints.index)]
    
    return spn_complaints, non_spn_complaints

@app.route('/')
def index():
    return render_template('front1.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'complaint_file' not in request.files:
        return "No file uploaded"
    
    file = request.files['complaint_file']
    if file.filename == '':
        return "No selected file"
    
    try:
        df = pd.read_excel(file, engine='openpyxl')  # Use openpyxl for better performance
        spn_df, non_spn_df = segregate_spn_nonspn(df)
        
        if spn_df is None and non_spn_df is None:
            return "Invalid file: Missing 'Observation' column"
        
        output_file = "Segregated_Complaints.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            spn_df.to_excel(writer, sheet_name='SPN_Complaints', index=False)
            non_spn_df.to_excel(writer, sheet_name='Non_SPN_Complaints', index=False)
        
        # Load the workbook to apply formatting
        wb = load_workbook(output_file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            
            # Apply Teal Green color to 'closed' and 'completed' in 'Incident Status' column
            teal_green_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
            
            if 'Incident Status' in df.columns:
                col_idx = df.columns.get_loc('Incident Status') + 1  # Excel column index (1-based)
                for row in range(2, ws.max_row + 1):  # Start from row 2 to skip header
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value and str(cell.value).strip().lower() in ['closed', 'completed']:
                        cell.fill = teal_green_fill
        
        wb.save(output_file)
        
        return send_file(output_file, as_attachment=True)
    except Exception as e:
        return f"Error processing file: {str(e)}"

if __name__ == '__main__':
    app.run(debug=True)