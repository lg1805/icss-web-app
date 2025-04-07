from flask import Flask, render_template, request, send_file
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

# Function to segregate SPN and Non-SPN complaints
def segregate_spn_nonspn(df):
    if 'Observation' not in df.columns:
        return None, None
    
    df['Observation'] = df['Observation'].astype(str)
    spn_complaints = df[df['Observation'].str.contains(r'\bspn\b', case=False, na=False, regex=True)]
    non_spn_complaints = df[~df.index.isin(spn_complaints.index)]
    
    return spn_complaints, non_spn_complaints

# Function to apply red color to "open" Incident Status
def apply_color_coding(file_path):
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Find the column index for Incident Status
        status_col_index = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == 'Incident Status':
                status_col_index = col
                break
        
        if status_col_index is None:
            continue
        
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_col_index)
            if str(cell.value).strip().lower() == 'open':
                cell.fill = red_fill
    
    wb.save(file_path)

@app.route('/')
def index():
    return render_template('front1.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'complaint_file' not in request.files:
        return "No file uploaded"
    
    file = request.files['complaint_file']
    if file.filename == '':
        return "No selected file"
    
    try:
        df = pd.read_excel(file, engine='openpyxl')
        spn_df, non_spn_df = segregate_spn_nonspn(df)
        
        if spn_df is None and non_spn_df is None:
            return "Invalid file: Missing 'Observation' column"
        
        output_file = "Segregated_Complaints.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            spn_df.to_excel(writer, sheet_name='SPN_Complaints', index=False)
            non_spn_df.to_excel(writer, sheet_name='Non_SPN_Complaints', index=False)
        
        apply_color_coding(output_file)
        
        return send_file(output_file, as_attachment=True)
    except Exception as e:
        return f"Error processing file: {str(e)}"

if __name__ == '__main__':
    app.run(debug=True)