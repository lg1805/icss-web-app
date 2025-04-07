from flask import Flask, request, render_template, send_file
import pandas as pd
import os
import joblib
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads/processed/'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

MODEL_PATH = r"D:\Lakshya\Project\ICSS_VSCode\PROJECT\ICSS Web App\random_forest_model.pkl"

def load_model():
    try:
        model_data = joblib.load(MODEL_PATH)
        model, vectorizer = model_data if isinstance(model_data, tuple) else (model_data, None)
        if not hasattr(model, 'predict'):
            raise ValueError("Loaded model does not have a 'predict' method")
        print("Model loaded successfully")
        return model, vectorizer
    except Exception as e:
        print(f"Error loading model: {e}")
        return None, None

model, vectorizer = load_model()

def calculate_rpn(severity, occurrence, detection):
    return severity * occurrence * detection

def predict_priority(observation):
    try:
        if model and vectorizer:
            transformed_text = vectorizer.transform([str(observation)])
            priority = model.predict(transformed_text)[0]
            
            severity_map = {'High': 10, 'Moderate': 5, 'Low': 2}
            occurrence_map = {'High': 9, 'Moderate': 6, 'Low': 3}
            detection_map = {'High': 2, 'Moderate': 5, 'Low': 8}
            
            severity = severity_map.get(priority, 1)
            occurrence = occurrence_map.get(priority, 5)
            detection = detection_map.get(priority, 4)
            
            rpn = calculate_rpn(severity, occurrence, detection)
            return priority, rpn
    except Exception as e:
        print(f"Prediction Error: {e}")
    return 'Low', 10

@app.route('/')
def index():
    return render_template('front1.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'complaint_file' not in request.files:
        return "No complaint_file part", 400
    
    file = request.files['complaint_file']
    if file.filename == '':
        return "No selected file", 400
    
    if file:
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        
        try:
            df = pd.read_excel(filepath)
        except Exception as e:
            return f"Error reading file: {e}", 400
        
        if 'Observation' not in df.columns or 'Incident Status' not in df.columns:
            return "Error: Required columns are missing in the uploaded file.", 400

        # Predict Priority and RPN
        df[['Priority', 'RPN']] = df['Observation'].fillna('').apply(lambda obs: pd.Series(predict_priority(obs)))

        # Separate SPN and Non-SPN complaints
        spn_df = df[df['Observation'].str.contains('spn', case=False, na=False)]
        non_spn_df = df[~df['Observation'].str.contains('spn', case=False, na=False)]

        # Save to separate sheets
        processed_filepath = os.path.join(UPLOAD_FOLDER, 'processed_' + file.filename)
        with pd.ExcelWriter(processed_filepath, engine='openpyxl') as writer:
            spn_df.to_excel(writer, sheet_name='SPN_Complaints', index=False)
            non_spn_df.to_excel(writer, sheet_name='Non_SPN_Complaints', index=False)

        # Apply color coding in Excel sheets
        wb = load_workbook(processed_filepath)
        
        for sheet_name in ['SPN_Complaints', 'Non_SPN_Complaints']:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            status = cell.value.lower()
                            if "closed" in status or "completed" in status:
                                cell.fill = green_fill
                            elif "pending" in status:
                                cell.fill = yellow_fill
                            else:
                                cell.fill = red_fill

        wb.save(processed_filepath)
        return send_file(processed_filepath, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)