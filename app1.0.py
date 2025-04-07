from flask import Flask, render_template, request, send_file, redirect, url_for, flash
import pandas as pd
import pickle
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Replace with a secure key

# Load your pre-trained vectorizer and classifier
with open('model/vectorizer.pkl', 'rb') as f:
    vectorizer = pickle.load(f)

with open('model/complaint_classifier.pkl', 'rb') as f:
    classifier = pickle.load(f)

# Define a mapping for classifier outputs to priority labels
priority_mapping = {0: 'Low', 1: 'Moderate', 2: 'High'}

def process_complaints(df):
    """
    Classify complaints using the 'Observation' column and add a 'Priority' column.
    """
    if 'Observation' not in df.columns:
        raise Exception("Column 'Observation' not found in the uploaded file.")
    
    texts = df['Observation'].astype(str).tolist()
    features = vectorizer.transform(texts)
    predictions = classifier.predict(features)
    priorities = [priority_mapping.get(pred, 'Unknown') for pred in predictions]
    df['Priority'] = priorities
    return df

def apply_highlighting(excel_buffer):
    """
    Open the generated Excel file and apply color formatting based on:
     - Creation Date: 
         * < 24 hours: Blue
         * 24 to 48 hours: Yellow
         * >= 48 hours: Red
     - Closed/completed complaints (based on 'Incident Status.') are highlighted in Green.
    
    This function assumes the worksheet has the columns:
     - "Incident Id"
     - "Creation Date"
     - "Incident Status."
    """
    wb = load_workbook(excel_buffer)
    ws = wb.active

    # Get header names (assumed to be in the first row)
    header = [cell.value for cell in ws[1]]
    try:
        incident_col = header.index("Incident Id") + 1
        creation_col = header.index("Creation Date") + 1
        status_col = header.index("Incident Status") + 1
    except ValueError:
        raise Exception("Required columns ('Incident Id', 'Creation Date', 'Incident Status') not found.")

    # Define fill colors
    fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")   # Blue
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red
    fill_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")   # Green

    now = datetime.now()

    # Process each data row (skip header)
    for row in ws.iter_rows(min_row=2):
        creation_cell = row[creation_col - 1]
        status_cell = row[status_col - 1]
        incident_cell = row[incident_col - 1]

        status = str(status_cell.value).strip().lower() if status_cell.value else ""
        if status in ["closed", "completed"]:
            incident_cell.fill = fill_green
            continue

        # Parse the creation date
        try:
            creation_date = pd.to_datetime(creation_cell.value)
        except Exception:
            continue

        diff_hours = (now - creation_date).total_seconds() / 3600.0

        if diff_hours < 24:
            incident_cell.fill = fill_blue
        elif diff_hours < 48:
            incident_cell.fill = fill_yellow
        else:
            incident_cell.fill = fill_red

    output_buffer = BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        try:
            # Read the uploaded Excel file
            df = pd.read_excel(file)
            
            # Process complaints: classify using 'Observation' and add 'Priority'
            df = process_complaints(df)
            
            # Write the processed DataFrame to an in-memory Excel file using a context manager
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Processed Complaints')
            output.seek(0)

            # Apply highlighting based on 'Creation Date' and 'Incident Status.'
            highlighted_file = apply_highlighting(output)

            return send_file(highlighted_file,
                             as_attachment=True,
                             download_name='segregated_complaints.xlsx',
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e:
            flash(str(e))
            return redirect(request.url)
    return render_template('index1.0.html')

if __name__ == '__main__':
    app.run(debug=True)